'''b = '['
c = '['
for i in range(200):
    a = input()
    if i % 4 == 1:
        b += '\'' + a + '\'' + ', '
    if i % 4 == 0:
        c += '\'' + a + '\'' + ', '
b = b[:-2] + ']'
c = c[:-2] + ']'
print(b)
print(c)'''
import pandas as pd
import Functions as Fx

import Arrays as Ar

usa_cities = ['Birmingham', 'Anchorage', 'Phoenix', 'Little Rock', 'Los Angeles', 'Denver', 'Bridgeport', 'Wilmington',
              'Jacksonville', 'Atlanta', 'Honolulu', 'Boise', 'Chicago', 'Indianapolis', 'Des Moines', 'Wichita',
              'Louisville', 'New Orleans', 'Portland', 'Baltimore', 'Boston', 'Detroit', 'Minneapolis', 'Jackson',
              'Kansas City', 'Billings', 'Omaha', 'Las Vegas', 'Manchester', 'Newark', 'Albuquerque', 'New York City',
              'Charlotte', 'Fargo', 'Columbus', 'Oklahoma City', 'Portland', 'Philadelphia', 'Providence', 'Charleston',
              'Sioux Falls', 'Memphis', 'Houston', 'Salt Lake City', 'Burlington', 'Virginia Beach', 'Seattle',
              'Charleston', 'Milwaukee', 'Cheyenne']
usa_states = ['Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut', 'Delaware',
              'Florida', 'Georgia', 'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana',
              'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri', 'Montana',
              'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 'North Carolina',
              'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island', 'South Carolina',
              'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 'West Virginia',
              'Wisconsin', 'Wyoming']
can_cities = ['Calgary', 'Vancouver', 'Winnipeg', 'Moncton', 'St. John\'s', 'Yellowknife', 'Halifax', 'Iqaluit',
              'Toronto', 'Charlottetown', 'Montreal', 'Saskatoon', 'Whitehorse']
can_provinces = ['Alberta', 'British Columbia', 'Manitoba', 'New Brunswick', 'Newfoundland and Labrador',
                 'Northwest Territories', 'Nova Scotia', 'Nunavut', 'Ontario', 'Prince Edward Island', 'Quebec',
                 'Saskatchewan', 'Yukon']
mex_regions = ['Mexico North West', 'Mexico North East', 'Mexico Interior West', 'Mexico Interior', 'Mexico South West']
mex_states = ['Baja California', '‎Chihuahua', 'Jalisco', 'Mexico City', 'Yucatan']
mex_cities = ['Tijuana', 'Juarez', 'Guadalajara', 'Mexico City', 'Merida']
all_cities = can_cities + mex_cities + usa_cities
all_states = can_provinces + mex_regions + usa_states
all_states_2 = ['AB', 'BC', 'MB', 'NB', 'NL', 'NT', 'NS', 'NU', 'ON', 'PE', 'QC', 'SK', 'YT', 'MNW', 'MNE', 'MIW',
                'MIN', 'MSW', 'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN',
                'IA', 'KS',
                'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ', 'NM', 'NY', 'NC',
                'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY',
                ]
all_dict = Ar.create_lookup_dict(all_states, all_states_2)
all_dict_reverse = Ar.create_lookup_dict(all_states_2, all_states)

'''import time
from geopy.geocoders import Nominatim
i = 0
asd = [[]]
for city, state in zip(can_cities + mex_cities + usa_cities, can_provinces + mex_regions + usa_states):
    i += 1
    m = ['hjdkfhj', 'sdfsdf', 'dsfsdfa', 'dfgdfgd', 'erwertr', 'hghfg', 'sdfhjkd', 'dfgsdfgs', 'erfadfadf', 'fgsfgvcb']
    geolocator = Nominatim(user_agent= m[i%10])
    location = geolocator.geocode(city + ', ' + state)
    time.sleep(1.5)
    print(location, location.latitude, location.longitude)
    asd += [[city, state, location.latitude, location.longitude]]
import pandas as pd
df = pd.DataFrame(asd)
df.to_csv('coords.csv')'''

from geopy.distance import vincenty

df = pd.read_csv('coords.csv', index_col=0, skiprows=0)
df.columns = ['City', 'State', 'Lat', 'Long']
df = df.set_index('State')
final_flows = []
for state1 in all_states:
    temp = []
    for state2 in all_states:
        dist = vincenty((df.loc[state1, 'Lat'], df.loc[state1, 'Long']),
                        (df.loc[state2, 'Lat'], df.loc[state2, 'Long'])).kilometers
        temp += [dist]
    final_flows += [temp]
df2 = pd.DataFrame(final_flows, columns=all_states_2)
df2 = df2.set_index(pd.Series(all_states_2))
df2.to_csv('final_dist.csv')

import numpy as np

pre_ff = np.array(final_flows)
post_ff = pre_ff * 0.0012872276031893294 + 0.42821891169049636
post_ff2 = post_ff * 1000
df3 = pd.DataFrame(post_ff2, columns=all_states_2)
df3 = df3.set_index(pd.Series(all_states_2))

C8 = df3
C10 = 50  # Technical Lifetime
C11 = 3  # Construction Time
C12 = 0.05  # Interest Rate
C13 = 0.10  # Decomission Share
C14 = 0.08  # Discount Rate
C16 = C8 * (1 + C13 * np.e ** (-C14 * (C10))) / (((1 + C12) ** C11 - 1) / (C12 * (1 + C12) ** C11)) * C11
df4 = C16 / (((1 + C14) ** C10 - 1) / (C14 * (1 + C14) ** C10))

df3.to_excel("Pipeline Costs.xlsx", startrow=7, startcol=1, sheet_name="Overnight Cost")
df4.to_excel("Pipeline Costs2.xlsx", startrow=7, startcol=1, sheet_name="Investment Cost")

import openpyxl
xfile1 = openpyxl.load_workbook('Pipeline Costs.xlsx')
xfile2 = openpyxl.load_workbook('Pipeline Costs2.xlsx')
sheet1 = xfile1.get_sheet_by_name('Overnight Cost')
sheet2 = xfile2.get_sheet_by_name('Investment Cost')
for index, name in enumerate(all_states):
    sheet1['A' + str(index + 9)] = name
    sheet1.cell(row=7, column=index + 3).value = name
    sheet1.cell(row=index + 9, column=1).value = name
    sheet2['A' + str(index + 9)] = name
    sheet2.cell(row=7, column=index + 3).value = name
    sheet2.cell(row=index + 9, column=1).value = name
sheet1["A1"] = "Overnight Cost for Pipeline Construction between States"
sheet1["A2"] = "Million$ / (BCF/Day)"
sheet1["A3"] = "2018 Base Price"
sheet2["A1"] = "Investment Cost for Pipeline Construction between States"
sheet2["A2"] = "(Million$/Year) / (BCF/Day)"
sheet2["A3"] = "2018 Base Price"
xfile1.save('Pipeline Costs.xlsx')
xfile2.save('Pipeline Costs2.xlsx')

